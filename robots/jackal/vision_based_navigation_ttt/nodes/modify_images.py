#!/usr/bin/env python3
from tkinter import W
import numpy as np
import os
import pandas as pd
import xlsxwriter
from xlsxwriter import Workbook
import openpyxl
# importing PIL Module
from PIL import Image as im
import cv2
import json

def get_variables():
    # print("get_variables")
    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt/"
    file = open(path + "variables.json")
    data = json.load(file)
    file.close()
    count = data["count"]
    count_2 = data["count_2"]
    return count, count_2
    
def update_variables(count, count_2):
    # print("update_variables")
    path = os.environ["HOME"] + "/catkin_ws/src/vision_based_navigation_ttt/"
    file = open(path + "variables.json", "w")
    updated_data = {"count": count, "count_2": count_2}
    json.dump(updated_data, file)
    file.close()

def save_image(count : int, shared_path, curr_image): 
        img_name= str(count) + '.png'
        path = shared_path + img_name
        curr_image = im.fromarray(curr_image)
        curr_image.save(path)

def flip_images():
    count, count_2 = get_variables()

    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt/"
    path_tau = path + "tau_values_real_data/tau_value"    # modify
    path_folder = path + "training_images_real_data/" #modify

    folders = [file for file in sorted(os.listdir(path_folder)) if os.path.isdir(os.path.join(path_folder, file))]
    for folder in folders:
        count_2 += 1
        path_images = path_folder + folder + '/'
        
        vel = folder.split('_')[4]
        path_new_folder = path_folder + 'training_images_' + str(count_2) + '_v_' + vel + '/'
        # create_folder
        os.mkdir(path_new_folder)
        images_in_folder = [f for f in sorted(os.listdir(path_images)) if f.endswith(".png")]

        print(images_in_folder,'imgfol')
        for idx in range(len(images_in_folder)-1) : 
            # print('ll', images_in_folder[idx])
            try:
                # load the original input image
                image = cv2.imread(path_images + images_in_folder[idx])

                # flip the original image horizontally
                horz_img = cv2.flip(image, 1)
                print("path",path)
                save_image(count, path_new_folder, horz_img)

                # retrive the direction from the filename
                ps = openpyxl.load_workbook(path_tau + str(images_in_folder[idx+1].split('.')[0]) + '.xlsx')
                sheet = ps['Sheet1']
                tau_val = [sheet['A1'].value,sheet['B1'].value,sheet['C1'].value,sheet['D1'].value,sheet['E1'].value]
                print(count ,tau_val)
                wb_tau =  xlsxwriter.Workbook(path_tau + str(count) + ".xlsx")
                wksheet_tau = wb_tau.add_worksheet()

                wksheet_tau.write('A1',tau_val[4])
                wksheet_tau.write('B1',tau_val[3])
                wksheet_tau.write('C1',tau_val[2])
                wksheet_tau.write('D1',tau_val[1])
                wksheet_tau.write('E1',tau_val[0])

                wb_tau.close()
                count += 1
                print('here')
                update_variables(count, count_2)

            except Exception as inst:
                print(idx)
                print(inst)


def change_brightness():
    count_new = get_variables()
    path_tau = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt/tau_values/tau_value"   
    path_folder = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt/training_images/"

    folders = [file for file in os.listdir(path_folder) if os.path.isdir(os.path.join(path_folder, file))]
    print('folders',folders)
    for folder in folders:
        print('folder',folder)
        # create_folder
        os.mkdir(path_folder + folder + "_brightness" )   
        path_images = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt/training_images/" + folder + '/'
        images_in_folder = [f for f in listdir(path_images) if f.endswith(".png")]

        for idx in range(len(images_in_folder)) : 
            print('ll', images_in_folder[idx])
            try:
                # load the original input image
                image = cv2.imread(path_images + images_in_folder[idx],0)

                # # Displaying the image
                # cv2.imshow('image', image)
                # # waits for user to press any key
                # cv2.waitKey(0)

                # flip the original image horizontally
                alphas = [0.5, 2]
                for alpha in alphas:

                    beta = 0
                    brightness_img = cv2.convertScaleAbs(image, alpha=alpha, beta=beta)
                    path = path_folder + folder + "_brightness/" + str(count_new) + ".png"
                    print("path",path)
                    cv2.imwrite(path, brightness_img)

                    # # Displaying the image
                    # cv2.imshow("brightness", brightness_img)
                    # # waits for user to press any key
                    # cv2.waitKey(0)

                    # retrive the direction from the filename
                    ps = openpyxl.load_workbook(path_tau + str(images_in_folder[idx+1].split('.')[0]) + '.xlsx')
                    sheet = ps['Sheet1']
                    tau_val = [sheet['A1'].value,sheet['B1'].value,sheet['C1'].value,sheet['D1'].value,sheet['E1'].value]
                    print(count_new ,tau_val)
                    wb_tau =  xlsxwriter.Workbook(path_tau + str(count_new) + ".xlsx")
                    wksheet_tau = wb_tau.add_worksheet()

                #    tau_val = [tau_el, tau_l, tau_c, tau_r, tau_er]
                    wksheet_tau.write('A1',tau_val[0])
                    wksheet_tau.write('B1',tau_val[1])
                    wksheet_tau.write('C1',tau_val[2])
                    wksheet_tau.write('D1',tau_val[3])
                    wksheet_tau.write('E1',tau_val[4])

                    wb_tau.close()
                    count_new += 1

            except Exception as inst:
                print(idx)
                print(inst)
    update_variables(count_new) 

def tau():
    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt/"
    path_tau = path + "tau_tr/" 

    files = [file for file in sorted(os.listdir(path_tau))]
    print(files)
    for file in files:
        # print(file.split('_')[1])
    # retrive the direction from the filename
        s=file.split('_')[1]
        # print(s)
        s=s[1:len(s)]
        # print(s)
        os.rename(os.path.join(path_tau,file), os.path.join(path_tau,'tau_'+ s))

# change_brightness()
flip_images()
# tau()
